import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from scipy.stats import norm
import io
from openpyxl import load_workbook # <--- 核心引入：用于保留格式

# ==========================================
# 1. 页面配置
# ==========================================
st.set_page_config(
    page_title="成绩分布交互分析台",
    page_icon="🎓",
    layout="wide"  
)

# 顶级审美配置 (Ivy Prestige Theme)
COLOR_BG = '#FFFFF0'
COLOR_TEXT = '#1C1C1C'
COLOR_MAIN = '#004225'
COLOR_ACCENT = '#D4AF37'


import matplotlib.font_manager as fm
import os


# ==========================================
# 字体设置 (云端部署专用版)
# ==========================================
# 假设您已经把 SimHei.ttf 文件上传到了同级目录
import matplotlib.font_manager as fm
import os

# ==========================================
# 字体设置 (GitHub 云端部署修正版)
# ==========================================
# 1. 动态获取当前脚本 app.py 所在的目录
current_dir = os.path.dirname(os.path.abspath(__file__))

# 2. 拼接字体路径 
# ⚠️ 关键修正：根据您的 GitHub 截图，文件名必须是全小写的 'simhei.ttf'
font_path = os.path.join(current_dir, 'simhei.ttf')

if os.path.exists(font_path):
    # 3. 核心步骤：强制将字体注册到 Matplotlib 管理器中
    # (这一步解决了 "findfont: Font family not found" 的报错)
    fm.fontManager.addfont(font_path)
    
    # 4. 获取该字体的内部注册名称 (防止它内部叫 'SimHei Regular' 而不是 'SimHei')
    font_prop = fm.FontProperties(fname=font_path)
    custom_font_name = font_prop.get_name()
    
    # 5. 设置为全局默认字体
    plt.rcParams['font.family'] = custom_font_name
    print(f"✅ 成功加载并注册本地字体: {custom_font_name}")
else:
    # 调试信息：如果找不到，打印出来方便排查
    print(f"⚠️ 未找到字体文件，请检查路径: {font_path}")
    # 回退方案
    plt.rcParams['font.sans-serif'] = ['Microsoft YaHei', 'SimHei', 'Arial', 'sans-serif']

# 解决负号显示为方块的问题
plt.rcParams['axes.unicode_minus'] = False


# ==========================================
# 2. 辅助函数
# ==========================================

def draw_chart(data, col_name=""):
    """绘制正态分布图"""
    mu, std = data.mean(), data.std()
    
    fig, ax = plt.subplots(figsize=(10, 6))
    
    # 直方图
    ax.hist(data, bins=15, density=True, color=COLOR_MAIN, alpha=0.2, 
            edgecolor=COLOR_MAIN, linewidth=1, rwidth=0.9)
    
    # 正态曲线
    xmin, xmax = data.min() - 5, data.max() + 5
    if xmax - xmin < 10: xmin -= 5; xmax += 5
    ax.set_xlim(xmin, xmax)
    x = np.linspace(xmin, xmax, 300)
    p = norm.pdf(x, mu, std)
    
    ax.plot(x, p, color=COLOR_MAIN, linewidth=3, label='理论正态分布')
    ax.fill_between(x, p, color=COLOR_MAIN, alpha=0.05)
    
    # 辅助线
    ax.axvline(mu, color=COLOR_ACCENT, linestyle='--', linewidth=2, label='平均分')
    
    # 统计信息
    stats_text = f'平均分 = {mu:.2f}\n标准差 = {std:.2f}'
    ax.text(0.05, 0.95, stats_text, transform=ax.transAxes, fontsize=12,
            verticalalignment='top', horizontalalignment='left',
            color=COLOR_MAIN,
            bbox=dict(boxstyle='round,pad=0.5', facecolor=COLOR_BG, edgecolor=COLOR_ACCENT, alpha=0.8))

    # 修饰
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.grid(axis='y', linestyle=':', alpha=0.4)
    ax.legend(frameon=False, loc='upper left', bbox_to_anchor=(0, 0.85), fontsize=10)
    
    title_str = f"{col_name} 分布概览" if col_name else "成绩分布概览"
    ax.set_title(title_str, fontsize=16, pad=15, color=COLOR_MAIN, fontweight='bold')
    fig.tight_layout()
    
    return fig, mu, std

def update_excel_formatting(df_new, original_file_obj):
    """
    核心黑科技：
    打开原始 Excel (保留格式)，将 df_new 的值填入，
    处理行数变化，最后返回二进制流。
    """
    # 1. 重置文件指针，确保从头读取
    original_file_obj.seek(0)
    
    # 2. 使用 openpyxl 加载原始文件 (keep_vba=False, data_only=False 以保留样式)
    wb = load_workbook(original_file_obj)
    ws = wb.active # 默认操作第一个 Sheet
    
    # 3. 将 DataFrame 转换为列表 (不包含表头，因为表头通常不动)
    # 注意：我们假设列的顺序没有变。如果用户拖拽了列序，这里需要更复杂的逻辑。
    # 这里我们只更新数据部分（从第2行开始）
    data_rows = df_new.values.tolist()
    
    # 4. 填入新数据 (保留单元格原有样式)
    # enumerate 从 0 开始，Excel 行从 2 开始 (1是表头)
    for row_idx, row_data in enumerate(data_rows):
        excel_row = row_idx + 2 
        for col_idx, value in enumerate(row_data):
            excel_col = col_idx + 1
            # 更新值，openpyxl 会自动保留该单元格的颜色/字体/边框
            ws.cell(row=excel_row, column=excel_col).value = value
            
    # 5. 处理行数删除的情况
    # 如果新数据比老数据少，需要把 Excel 里多余的老数据行删掉
    current_max_row = ws.max_row
    new_data_count = len(data_rows) + 1 # +1 是因为有表头
    
    if current_max_row > new_data_count:
        # 删除多余的行
        ws.delete_rows(new_data_count + 1, amount=(current_max_row - new_data_count))
        
    # 6. 保存到内存
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# ==========================================
# 3. 主界面逻辑
# ==========================================
st.title("🎓 成绩分布交互分析台")
st.markdown("---")

# --- 侧边栏：文件上传 ---
with st.sidebar:
    st.header("📂 数据导入")
    uploaded_file = st.file_uploader("上传 Excel 文件", type=["xlsx", "xls"])
    st.info("💡 提示：导出时将完美保留原表格的颜色、字体和格式。")

# --- 数据处理与布局 ---
if uploaded_file is not None:
    try:
        # 1. 读取上传的文件
        df_original = pd.read_excel(uploaded_file, engine='openpyxl')
        
        # 2. 列筛选逻辑
        exclude_cols = ['序号', '班级', '学号', '姓名', '备注', 'ID', 'No']
        score_cols = []
        for col in df_original.columns:
            if col in exclude_cols or any(x in str(col) for x in ['学号', '姓名', '班级']):
                continue
            try:
                sample = df_original[col].dropna()
                if len(sample) > 0:
                    pd.to_numeric(sample, errors='raise')
                    score_cols.append(col)
            except:
                pass
        
        if not score_cols:
            st.error("⚠️ 未在表格中找到可分析的【数字列】。")
        else:
            default_index = len(score_cols) - 1
            target_col = st.sidebar.selectbox("🎯 选择要分析的成绩列", score_cols, index=default_index)

            # ==========================================
            # 关键布局
            # ==========================================
            col1, col2 = st.columns([4, 6], gap="large")

            with col1:
                st.subheader("📝 数据编辑器")
                st.caption(f"当前正在编辑：**{target_col}**")
                
                # 数据编辑器
                df_edited = st.data_editor(
                    df_original, 
                    num_rows="dynamic",
                    height=600,
                    use_container_width=True,
                    key="data_editor"
                )

                # ======================================================
                # 【导出功能升级】使用格式保留逻辑
                # ======================================================
                # 调用我们写的 update_excel_formatting 函数
                # 传入：修改后的数据 + 原始文件对象
                final_buffer = update_excel_formatting(df_edited, uploaded_file)
                
                with st.sidebar:
                    st.markdown("---")
                    st.header("💾 导出结果")
                    st.download_button(
                        label="📥 下载 Excel (保留原格式)",
                        data=final_buffer,
                        file_name="成绩单_已更新.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        help="导出的文件将保持您上传时的所有格式（颜色、列宽等），仅更新数据。"
                    )

            with col2:
                st.subheader("📊 实时分布图")
                clean_data = pd.to_numeric(df_edited[target_col], errors='coerce').dropna()
                
                if len(clean_data) > 1:
                    fig, mu, std = draw_chart(clean_data, col_name=target_col)
                    st.pyplot(fig, use_container_width=True)
                    
                    c1, c2, c3 = st.columns(3)
                    c1.metric("平均分", f"{mu:.2f}")
                    c2.metric("标准差", f"{std:.2f}")
                    c3.metric("有效样本", f"{len(clean_data)}")
                else:
                    st.warning("⚠️ 有效数据太少，无法绘图。")

    except Exception as e:
        st.error(f"解析文件出错: {e}")

else:
    # --- 欢迎页 ---
    st.info("👋 请在左侧侧边栏上传 Excel 成绩单。")
    
    if st.button("或者：使用演示数据体验"):
        dummy_data = pd.DataFrame({
            '姓名': [f'学生{i}' for i in range(1, 51)],
            '平时成绩': np.random.randint(60, 100, 50),
            '期末考核(必填)': np.random.normal(75, 10, 50).astype(int)
        })
        
        col1, col2 = st.columns([4, 6], gap="large")
        with col1:
            st.subheader("📝 数据编辑器 (演示)")
            df_demo = st.data_editor(dummy_data, height=500, use_container_width=True)
            
            # 演示模式直接导出普通 Excel 即可
            buffer_demo = io.BytesIO()
            with pd.ExcelWriter(buffer_demo, engine='openpyxl') as writer:
                df_demo.to_excel(writer, index=False)
                
            with st.sidebar:
                st.markdown("---")
                st.header("💾 导出结果 (演示)")
                st.download_button(
                    label="📥 下载演示数据",
                    data=buffer_demo,
                    file_name="演示数据.xlsx"
                )
                
        with col2:
            st.subheader("📊 实时分布图 (演示)")
            d_clean = df_demo.iloc[:, -1]
            fig, mu, std = draw_chart(d_clean, col_name="期末考核(必填)")
            st.pyplot(fig, use_container_width=True)
